#!/usr/bin/env python3
"""Generate AVRAI Unified Budget spreadsheet.

Covers pre-seed (18 months) and seed (24 months) with:
  - Predictive salaries by role and stage
  - Personnel costs tied to Master Plan phases
  - One-time, recurring, and infrastructure costs
  - Phase milestones with "what this buys" and success criteria
  - Stage totals with base-case and range

Run:  python3 docs/business/generate_budget_xlsx.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Colors & Styles ──────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
HEADER_FONT = Font(name="Helvetica Neue", bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
SECTION_FONT = Font(name="Helvetica Neue", bold=True, size=12, color="1A1A2E")
TOTAL_FILL = PatternFill(start_color="C5CAE9", end_color="C5CAE9", fill_type="solid")
TOTAL_FONT = Font(name="Helvetica Neue", bold=True, size=11)
BODY_FONT = Font(name="Helvetica Neue", size=11)
TITLE_FONT = Font(name="Helvetica Neue", bold=True, size=16, color="1A1A2E")
SUBTITLE_FONT = Font(name="Helvetica Neue", size=11, color="666666")
NOTE_FONT = Font(name="Helvetica Neue", size=10, italic=True, color="555555")
THIN_BORDER = Border(bottom=Side(style="thin", color="CCCCCC"))
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")
CURRENCY_FMT = '"$"#,##0'
CURRENCY_FMT_2 = '"$"#,##0.00'


def style_header_row(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER


def style_section_row(ws, row, num_cols, label):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1)
    cell.value = label
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(vertical="center")


def style_total_row(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT


def write_headers(ws, row, headers):
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col).value = h
    style_header_row(ws, row, len(headers))
    return row + 1


def write_row(ws, row, values, money_cols=None, fmt=CURRENCY_FMT):
    if money_cols is None:
        money_cols = set()
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = val
        cell.font = BODY_FONT
        cell.border = THIN_BORDER
        if col in money_cols and isinstance(val, (int, float)):
            cell.number_format = fmt
        cell.alignment = WRAP
    return row + 1


def write_total(ws, row, values, money_cols=None, fmt=CURRENCY_FMT):
    if money_cols is None:
        money_cols = set()
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = val
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        if col in money_cols and isinstance(val, (int, float)):
            cell.number_format = fmt
        cell.alignment = WRAP
    return row + 1


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_note(ws, row, num_cols, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1)
    cell.value = text
    cell.font = NOTE_FONT
    cell.alignment = WRAP
    return row + 1


# ═══════════════════════════════════════════════════════════════════
# SHEET 1: SALARIES & PERSONNEL
# ═══════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Salaries & Personnel"
ws1.sheet_properties.tabColor = "1A1A2E"
set_col_widths(ws1, [30, 18, 18, 18, 18, 14, 14, 32])

ws1.merge_cells("A1:H1")
ws1.cell(row=1, column=1).value = "AVRAI — Salaries & Personnel"
ws1.cell(row=1, column=1).font = TITLE_FONT

ws1.merge_cells("A2:H2")
ws1.cell(row=2, column=1).value = (
    "February 2026  •  Pre-Seed (18 mo) + Seed (24 mo)  •  "
    "US market, fully loaded (base + tax + benefits/contractor markup)"
)
ws1.cell(row=2, column=1).font = SUBTITLE_FONT

row = 4

# ── Salary Guide ─────────────────────────────────────────────────
style_section_row(ws1, row, 8, "PREDICTIVE SALARY GUIDE — Annual Equivalent by Role & Stage")
row += 1
row = write_headers(ws1, row, [
    "Role", "Pre-Seed Low", "Pre-Seed High", "Seed Low", "Seed High",
    "Pre-Seed FTE", "Seed FTE", "Primary Phases"
])

salary_data = [
    ["Founder / Technical Lead",      0,      60000,  80000,  120000, 1.0, 1.0, "All"],
    ["ML Engineer",                    120000, 160000, 140000, 180000, 0.5, 1.0, "3, 4, 5, 8, 11"],
    ["Flutter / Mobile Engineer",      100000, 140000, 110000, 150000, 0.5, 1.0, "6, 7, 8, 10"],
    ["Backend / Infra Engineer",       110000, 150000, 120000, 160000, 0.5, 1.0, "6, 7, 8, 9, 12, 13"],
    ["Platform / Backend Engineer",    None,   None,   120000, 165000, None, 1.0, "11, 12, 13"],
    ["Product (Fractional)",           None,   None,   80000,  120000, None, 0.5, "8, 9, 10"],
    ["Design / UX",                    None,   None,   60000,  120000, None, 0.5, "Pre-beta, 12, 13"],
    ["Compliance / Legal (Retainer)",  None,   None,   None,   None,   None, None, "2, 9"],
]

for d in salary_data:
    row = write_row(ws1, row, d, money_cols={2, 3, 4, 5})

row += 1
row = write_note(ws1, row, 8,
    "Notes: Pre-seed salaries are annualized equivalents (actual spend depends on FTE and months active). "
    "Founder takes $0 until funded, then $48K/yr draw. Seed founder salary adjusts to $80-120K. "
    "Design is one-time contract ($10-30K) in pre-seed; contract or 0.5 FTE in seed. "
    "Compliance is retainer: $2-6K total pre-seed, $10-25K/yr seed."
)

row += 2

# ── Pre-Seed Personnel ────────────────────────────────────────────
style_section_row(ws1, row, 8, "PRE-SEED PERSONNEL — 18 Months")
row += 1
row = write_headers(ws1, row, [
    "Role", "FTE", "Months Active", "Start Month", "Annual Equiv.",
    "Base Case", "Low", "High"
])

preseed_personnel = [
    ["Founder / Technical Lead",  1.0,  18, 1,  "See note",  0,      0,      90000],
    ["ML Engineer",               0.5,  15, 4,  140000,      87500,  75000,  100000],
    ["Flutter / Mobile Engineer", 0.5,  11, 8,  120000,      55000,  45833,  64167],
    ["Backend / Infra Engineer",  0.5,  11, 8,  130000,      59583,  50417,  68750],
    ["Design / UX (one-time)",    "—",  "—", "—", "—",       20000,  10000,  30000],
    ["Compliance (retainer)",     "—",  18, 1,  "—",         4000,   2000,   6000],
]

for d in preseed_personnel:
    row = write_row(ws1, row, d, money_cols={5, 6, 7, 8})

row = write_total(ws1, row, [
    "PRE-SEED PERSONNEL TOTAL", "", "", "", "",
    226083, 183250, 358917
], money_cols={6, 7, 8})

row += 1
row = write_note(ws1, row, 8,
    "Base case: Founder at $0 draw, ML/mobile/backend at midpoint salaries, "
    "design at $20K one-time, compliance at $4K retainer. "
    "Founder draws $0 until raise; takes $48K/yr post-raise (below market to extend runway)."
)

row += 2

# ── Seed Personnel ────────────────────────────────────────────────
style_section_row(ws1, row, 8, "SEED PERSONNEL — 24 Months")
row += 1
row = write_headers(ws1, row, [
    "Role", "FTE", "Months Active", "Start Month", "Annual Salary",
    "Base Case", "Low", "High"
])

seed_personnel = [
    ["Founder / Technical Lead",     1.0, 24, 19, 100000,  200000, 160000, 240000],
    ["ML Engineer",                  1.0, 24, 19, 160000,  320000, 280000, 360000],
    ["Flutter / Mobile Engineer",    1.0, 24, 19, 130000,  260000, 220000, 300000],
    ["Backend / Infra Engineer",     1.0, 24, 19, 140000,  280000, 240000, 320000],
    ["Platform / Backend Engineer",  1.0, 24, 19, 142500,  285000, 240000, 330000],
    ["Product (Fractional)",         0.5, 24, 19, 100000,  100000, 80000,  120000],
    ["Design / UX (contract / 0.5)", 0.5, 24, 19, 90000,   90000,  60000,  120000],
    ["Compliance (retainer)",        "—", 24, 19, "—",      17500,  10000,  25000],
]

for d in seed_personnel:
    row = write_row(ws1, row, d, money_cols={5, 6, 7, 8})

row = write_total(ws1, row, [
    "SEED PERSONNEL TOTAL", "", "", "", "",
    1552500, 1290000, 1815000
], money_cols={6, 7, 8})

row += 1
row = write_note(ws1, row, 8,
    "Base case: All roles at midpoint salary. Product at 0.5 FTE. "
    "Design as ongoing contract or 0.5 FTE. Compliance retainer at $17.5K/yr."
)


# ═══════════════════════════════════════════════════════════════════
# SHEET 2: ALL COSTS
# ═══════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("All Costs")
ws2.sheet_properties.tabColor = "283593"
set_col_widths(ws2, [40, 18, 18, 18, 14, 50])

ws2.merge_cells("A1:F1")
ws2.cell(row=1, column=1).value = "AVRAI — Complete Cost Breakdown"
ws2.cell(row=1, column=1).font = TITLE_FONT

ws2.merge_cells("A2:F2")
ws2.cell(row=2, column=1).value = "All line items by category  •  Base case + range  •  Tied to Master Plan phases"
ws2.cell(row=2, column=1).font = SUBTITLE_FONT

row = 4

# ── One-Time: Legal & Formation ──────────────────────────────────
style_section_row(ws2, row, 6, "ONE-TIME — Legal & Formation")
row += 1
row = write_headers(ws2, row, ["Item", "Base Case", "Low", "High", "When", "Notes"])

onetime_legal = [
    ["LLC filings (AL + DE)",           490,    290,    490,    "Month 1",    "AL $200 + DE $90 + agent fees"],
    ["Operating agreements (2)",        1500,   1000,   3000,   "Month 1",    "Attorney-drafted for both LLCs"],
    ["EIN registration (2)",            0,      0,      0,      "Month 1",    "Free from IRS"],
    ["Business license (AL)",           150,    50,     300,    "Month 1",    "Varies by city/county"],
    ["Privacy policy (attorney)",       1500,   1000,   4000,   "By Phase 2", "Must match GDPR/CCPA implementation"],
    ["Terms of service (attorney)",     1500,   1000,   4000,   "By Phase 2", "Already linked from onboarding"],
    ["GDPR compliance review",          1500,   0,      3000,   "Phase 2",    "Only if targeting EU at launch"],
    ["D&O / general liability insurance", 2000, 1000,   3000,   "Month 1",    "Annual; required before investor money"],
]

for d in onetime_legal:
    row = write_row(ws2, row, d, money_cols={2, 3, 4})

row = write_total(ws2, row, ["SUBTOTAL: Legal & Formation", 8640, 4340, 17790, "", ""], money_cols={2, 3, 4})

row += 1

# ── One-Time: Intellectual Property ───────────────────────────────
style_section_row(ws2, row, 6, "ONE-TIME — Intellectual Property")
row += 1
row = write_headers(ws2, row, ["Item", "Base Case", "Low", "High", "When", "Notes"])

onetime_ip = [
    ["Patent attorney consultation",       3500,  2000,  5000,   "Month 2-3",   "Strategy + prior art search"],
    ["Prov. #1: AI2AI Personality Learning", 3500, 2000,  5000,  "Month 2-3",   "Core tech — most defensible IP"],
    ["Prov. #2: Quantum-Inspired Matching",  3500, 2000,  5000,  "Month 2-3",   "Quantum state + inner product matching"],
    ["Prov. #3: Offline-First Mesh Network", 3500, 2000,  5000,  "When fundable", "BLE mesh for AI2AI without internet"],
    ["Prov. #4: Knot Theory Group Dynamics", 3500, 2000,  5000,  "When fundable", "Topological knot invariant group formation"],
    ["Prov. #5: Privacy-Preserving Arch.",   3500, 2000,  5000,  "When fundable", "DP, k-anonymity, internal/external boundary"],
    ["Non-provisional conversion (each)",    15000, 7040, 23080, "Post-seed",    "12-month window from provisional. All 5 = $35-115K."],
]

for d in onetime_ip:
    row = write_row(ws2, row, d, money_cols={2, 3, 4})

row = write_total(ws2, row, ["SUBTOTAL: IP (provisionals only)", 21000, 12000, 30000, "", ""], money_cols={2, 3, 4})

row += 1

# ── One-Time: Design & Branding ───────────────────────────────────
style_section_row(ws2, row, 6, "ONE-TIME — Design & Branding")
row += 1
row = write_headers(ws2, row, ["Item", "Base Case", "Low", "High", "When", "Notes"])

onetime_design = [
    ["UI/UX design system + core flows", 7000,  3000,  12000, "Pre-beta (Phase 6-7)", "All screens, components, interactions, a11y"],
    ["Logo + app icon",                  1000,  200,   2000,  "With design",          "Must work 1024x1024 down to 29x29"],
]

for d in onetime_design:
    row = write_row(ws2, row, d, money_cols={2, 3, 4})

row = write_total(ws2, row, ["SUBTOTAL: Design", 8000, 3200, 14000, "", ""], money_cols={2, 3, 4})

row += 1

# ── One-Time: Security & Compliance ───────────────────────────────
style_section_row(ws2, row, 6, "ONE-TIME — Security & Compliance")
row += 1
row = write_headers(ws2, row, ["Item", "Base Case", "Low", "High", "When", "Notes"])

onetime_security = [
    ["Cryptographic security audit",  22000, 15000, 30000, "Pre-production (seed)", "Signal Protocol, AI2AI, BLE, pen test. Beta uses synthetic data only; audit before production launch."],
    ["Penetration test (app + API)",  "Incl.", "Incl.", "Incl.", "With audit", "Bundled in crypto audit above"],
]

for d in onetime_security:
    row = write_row(ws2, row, d, money_cols={2, 3, 4})

row = write_total(ws2, row, ["SUBTOTAL: Security", 22000, 15000, 30000, "", ""], money_cols={2, 3, 4})

row += 1

# ── One-Time: Hardware ────────────────────────────────────────────
style_section_row(ws2, row, 6, "ONE-TIME — Hardware & Devices")
row += 1
row = write_headers(ws2, row, ["Item", "Base Case", "Low", "High", "When", "Notes"])

onetime_hw = [
    ["MacBook Pro (M3/M4)",            2500,  1999,  3499,  "If needed",    "Required for iOS builds. Skip if owned."],
    ["Mac Mini (CI)",                   0,     0,     1299,  "If needed",    "Skip if using GitHub Actions"],
    ["External monitor (27\" 4K)",      0,     0,     799,   "If needed",    "Skip if owned"],
    ["External SSD + cables/dock",      200,   128,   649,   "If needed",    "Backups, ML files, Xcode archives"],
    ["iPhone 15 Pro (primary iOS)",     1099,  999,   1199,  "Before beta",  "Tier 3 test device"],
    ["iPhone 14 (compat. testing)",     650,   500,   799,   "Before beta",  "Tier 2 testing. Refurb OK."],
    ["Google Pixel 8 (primary Android)", 625,  549,   699,   "Before beta",  "Stock Android baseline"],
    ["Samsung Galaxy S24",              800,   699,   899,   "Before beta",  "Samsung One UI (BLE, notifs differ)"],
    ["Budget Android (~$150)",          200,   149,   299,   "Before beta",  "Tier 0-1 performance. Real users have cheap phones."],
]

for d in onetime_hw:
    row = write_row(ws2, row, d, money_cols={2, 3, 4})

row = write_total(ws2, row, ["SUBTOTAL: Hardware", 6074, 5024, 10141, "", "Skip Mac if owned = save ~$2.5K"], money_cols={2, 3, 4})

row += 1

# ── One-Time: Beta & User Acquisition ─────────────────────────────
style_section_row(ws2, row, 6, "ONE-TIME — Beta Program & User Acquisition")
row += 1
row = write_headers(ws2, row, ["Item", "Base Case", "Low", "High", "When", "Notes"])

onetime_beta = [
    ["Beta user recruitment",           1000,  500,   2000,  "Phase 10.6",  "Flyers, campus outreach, local community posts"],
    ["Beta incentives / gift cards",    500,   0,     1500,  "Phase 10.6",  "Thank-you for testers; not pay-for-installs"],
    ["Analytics instrumentation",       0,     0,     500,   "Phase 10.6",  "Firebase free; PostHog free tier; Mixpanel if needed"],
    ["Landing page (avrai.org)",        500,   0,     1500,  "Month 6+",    "Simple waitlist / pitch page"],
]

for d in onetime_beta:
    row = write_row(ws2, row, d, money_cols={2, 3, 4})

row = write_total(ws2, row, ["SUBTOTAL: Beta & Acquisition", 2000, 500, 5500, "", ""], money_cols={2, 3, 4})

row += 1

# ── Recurring: Tools & Ops ────────────────────────────────────────
style_section_row(ws2, row, 6, "RECURRING — Tools & Operations (Monthly)")
row += 1
row = write_headers(ws2, row, ["Item", "Pre-Seed /mo", "Seed /mo", "", "", "Notes"])

recurring_tools = [
    ["Supabase Pro",                    25,    200,   None, None, "Scales with users: $25 → $80 → $400"],
    ["Cursor Ultra + extra credits",    3700,  1500,  None, None, "Replaces ~$120K+ in eng salary. Drops with team."],
    ["GitHub (team plan)",              4,     40,    None, None, "$4/mo solo → $4/user/mo with team"],
    ["Google Workspace",                7,     20,    None, None, "@avrai.org email"],
    ["Apple Developer (amortized)",     8,     8,     None, None, "$99/yr"],
    ["Claude Pro / AI ops",             20,    50,    None, None, "Admin, legal research, drafting"],
    ["Accounting (Wave → QuickBooks)",  0,     30,    None, None, "Free until real transactions"],
    ["Slack / Linear / Notion",         0,     100,   None, None, "Free solo; paid with team"],
    ["Google Play (amortized)",         2,     2,     None, None, "$25 one-time, amortized"],
]

for d in recurring_tools:
    row = write_row(ws2, row, d, money_cols={2, 3})

row = write_total(ws2, row, ["SUBTOTAL: Tools /mo", 3766, 1950, None, None, ""], money_cols={2, 3})

row += 1

# ── Recurring: Infrastructure ─────────────────────────────────────
style_section_row(ws2, row, 6, "RECURRING — Infrastructure (Monthly)")
row += 1
row = write_headers(ws2, row, ["Item", "Pre-Seed /mo", "Seed /mo", "", "", "Notes / Phase trigger"])

recurring_infra = [
    ["Edge function overage",      0,     200,   None, None, "Beyond Supabase Pro 2M invocations"],
    ["Object storage (models)",    0,     50,    None, None, "Phase 5+ ONNX, Phase 6.7 SLM delivery"],
    ["Cloudflare R2 (CDN / OTA)",  0,     50,    None, None, "Free egress. Phase 5-6 model distribution."],
    ["ML training compute (GPU)",  50,    400,   None, None, "Colab free → RunPod/Lambda. Phase 4-5, 8."],
    ["Federation aggregation",     0,     500,   None, None, "Phase 8.1: federated gradient workers"],
    ["Locality / global state DB", 0,     200,   None, None, "Phase 8.9, 9.4: locality + tax rules"],
    ["Monitoring / error tracking", 0,    50,    None, None, "Sentry free tier → paid with users"],
    ["APIs (Maps, Places, etc.)",  0,     200,   None, None, "Free tier covers dev; paid at scale"],
]

for d in recurring_infra:
    row = write_row(ws2, row, d, money_cols={2, 3})

row = write_total(ws2, row, ["SUBTOTAL: Infrastructure /mo", 50, 1650, None, None, ""], money_cols={2, 3})

row += 1

# ── Recurring: Legal / Compliance ─────────────────────────────────
style_section_row(ws2, row, 6, "RECURRING — Legal & Compliance (Annual)")
row += 1
row = write_headers(ws2, row, ["Item", "Pre-Seed /yr", "Seed /yr", "", "", "Notes"])

recurring_legal = [
    ["DE registered agent",         150,   150,   None, None, "Required annually for DE LLC"],
    ["D&O / liability insurance",   2000,  3000,  None, None, "Scales modestly with revenue/headcount"],
    ["Ongoing legal consultation",  2000,  8000,  None, None, "Contract review, NDAs, employment. Quarterly."],
    ["Domain renewal (avrai.org)",  12,    12,    None, None, "Already owned"],
]

for d in recurring_legal:
    row = write_row(ws2, row, d, money_cols={2, 3})

row = write_total(ws2, row, ["SUBTOTAL: Legal /yr", 4162, 11162, None, None, ""], money_cols={2, 3})


# ═══════════════════════════════════════════════════════════════════
# SHEET 3: MILESTONES
# ═══════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Milestones")
ws3.sheet_properties.tabColor = "00897B"
set_col_widths(ws3, [10, 40, 10, 12, 48, 40])

ws3.merge_cells("A1:F1")
ws3.cell(row=1, column=1).value = "AVRAI — Phase Milestones & What Each Stage Buys"
ws3.cell(row=1, column=1).font = TITLE_FONT

ws3.merge_cells("A2:F2")
ws3.cell(row=2, column=1).value = "Master Plan phases mapped to deliverables, timelines, and success criteria"
ws3.cell(row=2, column=1).font = SUBTITLE_FONT

row = 4
row = write_headers(ws3, row, [
    "Phase", "Name", "Tier", "Duration", "What This Buys", "Success Criteria"
])

milestones = [
    ["1",  "Outcome Data & Memory Systems",        "Tier 0", "4-5 wks",
     "Episodic memory, outcome tuples, conviction memory — the data the entire world model trains on",
     "Episodic tuples recording for all user actions; nightly consolidation running"],
    ["2",  "Privacy, Compliance & Legal Infra",     "Tier 0", "3-4 wks",
     "GDPR-ready, privacy architecture, consent management — legal foundation for all user data",
     "Privacy audit passing; DPA templates ready; consent flow live"],
    ["3",  "World Model State & Action Encoders",   "Tier 1", "5-6 wks",
     "The system's \"senses\" — 166-176D feature extraction from 25+ services",
     "State encoder < 20ms; feature freshness tracking live; list quantum entity working"],
    ["4",  "Energy Function & Formula Replacement", "Tier 1", "6-8 wks",
     "Replaces 30+ hardcoded formulas with learned energy functions (VICReg). The brain's critic.",
     "Energy function < 10ms; A/B test framework live; first 5 formulas replaced"],
    ["5",  "Transition Predictor & On-Device Training", "Tier 1", "5-6 wks",
     "Predicts how users evolve over time. First on-device training loop.",
     "Transition predictor ONNX exported; on-device training in overnight window"],
    ["6",  "MPC Planner, System 1/2, SLM & Agent",  "Tier 2", "8-10 wks",
     "The decision-making brain: plans action sequences, instant reactions, language interface",
     "MPC scoring < 200ms for 50 candidates; SLM running on-device; agent operational"],
    ["7",  "Orchestrators, Life Pattern, Self-Optimization", "Tier 2", "14-18 wks",
     "The system runs itself: triggers, device tiers, self-healing, meta-learning, agent cognition",
     "Battery-adaptive scheduling; self-optimization experiments running; agent cognition live"],
    ["8",  "Ecosystem Intelligence & AI2AI",         "Tier 3", "10-13 wks",
     "AI agents talk to each other. Group matching. Federated learning. Behavioral archetypes.",
     "AI2AI mesh functional; federated gradients syncing; group formation working"],
    ["9",  "Business Ops & Marketplace",             "Parallel", "12-18 wks",
     "Revenue engine: business partnerships, marketplace, tax compliance, expertise system",
     "First business onboarded; payment flow live; tax doc generation working"],
    ["10", "Feature Completion & Polish",            "Parallel", "10-14 wks",
     "FIRST TESTABLE BUILD. Stubs replaced, friend system, analytics, codebase cleaned.",
     "10.6 = feature-complete testable build; zero critical bugs; onboarding flow polished"],
    ["11", "Industry Integrations & JEPA",           "Tier 3", "12-20 wks",
     "External data sources, JEPA research track, platform expansion",
     "2+ industry integrations live; JEPA prototype showing improvement over baseline"],
    ["12", "Admin Platform & Conviction Oracle",     "Tier 3", "14-20 wks",
     "Desktop admin, AI Code Studio, Partner SDK, value intelligence, universe conviction server",
     "Admin dashboard live; Conviction Oracle operational; Partner SDK documented"],
    ["13", "White-Label Federation & Universe Model", "Tier 3", "14-20 wks",
     "Fractal federation: organizations run their own AVRAI universes",
     "First org universe deployed; seamless adoption working; dual identity functional"],
    ["14", "Researcher Access Pathway",              "Tier 3", "4-6 wks",
     "IRB-compatible anonymized research data pipeline",
     "Research API live; k-anonymity verified; first research dataset exportable"],
]

for d in milestones:
    row = write_row(ws3, row, d)

row += 1
row = write_note(ws3, row, 6,
    "Pre-seed target: Through Phase 7, start Phases 8/9/10, reach Phase 10.6 (first testable build).  "
    "Seed target: Complete all phases through 14. Product and business that can scale."
)


# ═══════════════════════════════════════════════════════════════════
# SHEET 4: SUMMARY
# ═══════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Summary")
ws4.sheet_properties.tabColor = "4CAF50"
set_col_widths(ws4, [36, 20, 20, 20, 44])

ws4.merge_cells("A1:E1")
ws4.cell(row=1, column=1).value = "AVRAI — Budget Summary"
ws4.cell(row=1, column=1).font = TITLE_FONT

ws4.merge_cells("A2:E2")
ws4.cell(row=2, column=1).value = "Stage totals, monthly burn, and raise targets"
ws4.cell(row=2, column=1).font = SUBTITLE_FONT

row = 4

# ── Pre-Seed Summary ──────────────────────────────────────────────
style_section_row(ws4, row, 5, "PRE-SEED — 18 Months (Through Phase 7 + Start 8/9/10)")
row += 1
row = write_headers(ws4, row, ["Category", "Base Case", "Low", "High", "Notes"])

preseed_summary = [
    ["Personnel (18 mo)",          226083,  183250,  358917, "Founder $0 + ML/mobile/backend fractional + design + compliance"],
    ["One-time: Legal & formation", 8640,   4340,    17790,  "LLCs, legal docs, insurance"],
    ["One-time: IP (provisionals)", 21000,  12000,   30000,  "5 provisionals (defer 3-5 if tight)"],
    ["One-time: Design & branding", 8000,   3200,    14000,  "UI/UX system + logo"],
    ["One-time: Hardware & devices", 6074,  5024,    10141,  "Test devices; skip Mac if owned"],
    ["One-time: Beta program",      2000,   500,     5500,   "User recruitment, landing page"],
    ["Recurring: Tools (18 mo)",    67788,  4536,    70020,  "Cursor heavy at high; drops with team"],
    ["Recurring: Infra (18 mo)",    900,    0,       14400,  "$50/mo avg → $800/mo by month 18"],
    ["Recurring: Legal (18 mo)",    6243,   3000,    12000,  "Agent, insurance, consultation (1.5 yr)"],
]

for d in preseed_summary:
    row = write_row(ws4, row, d, money_cols={2, 3, 4})

row = write_total(ws4, row, [
    "SUBTOTAL", 346728, 215850, 532768, ""
], money_cols={2, 3, 4})

row = write_total(ws4, row, [
    "CONTINGENCY (15%)", 52009, 32378, 79915, ""
], money_cols={2, 3, 4})

row = write_total(ws4, row, [
    "PRE-SEED TOTAL", 398737, 248228, 612683, ""
], money_cols={2, 3, 4})

row += 1
row = write_note(ws4, row, 5,
    "Base case monthly burn: ~$19K/mo (personnel $12.6K + tools $3.8K + infra/legal $2.8K). "
    "Recommended pre-seed raise: $400K-$600K for 18 months to Phase 10.6 (first testable build)."
)

row += 2

# ── Seed Summary ──────────────────────────────────────────────────
style_section_row(ws4, row, 5, "SEED — 24 Months (Through Phase 14, Scalable Product)")
row += 1
row = write_headers(ws4, row, ["Category", "Base Case", "Low", "High", "Notes"])

seed_summary = [
    ["Personnel (24 mo)",            1552500, 1290000, 1815000, "4-5 FTE + product + design + compliance"],
    ["One-time: Security audit",     22000,   15000,   30000,   "Crypto, BLE, pen test — before production"],
    ["One-time: Non-provisional (2)", 30000,  14080,   46160,   "Convert top 2 patents; rest later"],
    ["One-time: Ongoing legal/compliance", 5000, 2000,  10000,  "Employment docs, vendor contracts"],
    ["Recurring: Tools (24 mo)",     46800,   9600,    72000,   "Cursor decreases; Slack/Linear increase"],
    ["Recurring: Infra (24 mo)",     39600,   12000,   72000,   "Federation, locality, ML, OTA"],
    ["Recurring: Legal (24 mo)",     22324,   8000,    40000,   "Agent, insurance, consultation (2 yr)"],
]

for d in seed_summary:
    row = write_row(ws4, row, d, money_cols={2, 3, 4})

row = write_total(ws4, row, [
    "SUBTOTAL", 1718224, 1350680, 2085160, ""
], money_cols={2, 3, 4})

row = write_total(ws4, row, [
    "CONTINGENCY (15%)", 257734, 202602, 312774, ""
], money_cols={2, 3, 4})

row = write_total(ws4, row, [
    "SEED TOTAL", 1975958, 1553282, 2397934, ""
], money_cols={2, 3, 4})

row += 1
row = write_note(ws4, row, 5,
    "Base case monthly burn: ~$72K/mo (personnel $64.7K + tools $2K + infra $1.7K + legal $3.5K). "
    "Recommended seed raise: $2M-$3M for 24 months to Phase 14 (product + business + platform). "
    "To compete with well-funded peers ($8-10M raises): plan $3M-$5M seed."
)

row += 2

# ── Raise Targets ─────────────────────────────────────────────────
style_section_row(ws4, row, 5, "RECOMMENDED RAISE TARGETS")
row += 1
row = write_headers(ws4, row, ["Stage", "Lean", "Base Case", "Compete", "What You Get"])

raise_targets = [
    ["Pre-Seed", 250000, 400000, 600000,
     "18 mo → Phase 10.6 (testable build, world model live, first beta)"],
    ["Seed", 1600000, 2000000, 3500000,
     "24 mo → Phase 14 (admin, federation, marketplace, research API)"],
    ["Combined (42 mo)", 1850000, 2400000, 4100000,
     "Full Master Plan: intelligence-first product + scalable business"],
]

for d in raise_targets:
    row = write_row(ws4, row, d, money_cols={2, 3, 4})

row += 1
row = write_note(ws4, row, 5,
    "\"Lean\" = founder at $0, minimal team, defer patents and audit. "
    "\"Base Case\" = midpoint salaries, full provisionals, design, audit at seed. "
    "\"Compete\" = full team earlier, all IP, buffer for pivots and market response."
)


# ── Save ──────────────────────────────────────────────────────────
output_path = "/Users/reisgordon/AVRAI/docs/business/AVRAI_BUDGET.xlsx"
wb.save(output_path)
print(f"\n✅ Saved to {output_path}")
print("   Open in Excel, Numbers, or Google Sheets.")
print("\n   Sheets:")
print("   1. Salaries & Personnel — Predictive salaries by role and stage")
print("   2. All Costs — Every line item (one-time + recurring)")
print("   3. Milestones — Phase timeline with \"what this buys\"")
print("   4. Summary — Stage totals, burn, raise targets")
