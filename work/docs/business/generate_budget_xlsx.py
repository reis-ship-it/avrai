#!/usr/bin/env python3
"""Generate AVRAI Budget spreadsheet for Apple Numbers / Excel.

Solo founder, pre-pre-seed budget.
Two LLCs: RG Enterprises (Alabama) and AVRAI (Delaware).
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Colors & Styles ──────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
HEADER_FONT = Font(name="Helvetica Neue", bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
SECTION_FONT = Font(name="Helvetica Neue", bold=True, size=12, color="1A1A2E")
TOTAL_FILL = PatternFill(start_color="C5CAE9", end_color="C5CAE9", fill_type="solid")
TOTAL_FONT = Font(name="Helvetica Neue", bold=True, size=11)
BODY_FONT = Font(name="Helvetica Neue", size=11)
TITLE_FONT = Font(name="Helvetica Neue", bold=True, size=16, color="1A1A2E")
SUBTITLE_FONT = Font(name="Helvetica Neue", size=11, color="666666")
THIN_BORDER = Border(bottom=Side(style="thin", color="CCCCCC"))
WRAP = Alignment(wrap_text=True, vertical="top")
CURRENCY_FMT = '"$"#,##0'
CURRENCY_FMT_2 = '"$"#,##0.00'


def style_header_row(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")


def style_section_row(ws, row, num_cols, label):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1)
    cell.value = label
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(vertical="center")


def style_total_row(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT


def write_headers(ws, row, headers):
    num_cols = len(headers)
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col).value = h
    style_header_row(ws, row, num_cols)
    return row + 1


def write_row(ws, row, values, money_cols=None, fmt=CURRENCY_FMT):
    """Write a data row. money_cols = set of 1-based column indices to format as currency."""
    if money_cols is None:
        money_cols = set()
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = val
        cell.font = BODY_FONT
        cell.border = THIN_BORDER
        if col in money_cols and isinstance(val, (int, float)):
            cell.number_format = fmt
        cell.alignment = WRAP
    return row + 1


def write_total_row(ws, row, label, num_cols, sum_cols=None, data_start=None, data_end=None, fmt=CURRENCY_FMT):
    """Write a styled total row with SUM formulas."""
    ws.cell(row=row, column=1).value = label
    if sum_cols and data_start and data_end:
        for col in sum_cols:
            col_letter = get_column_letter(col)
            cell = ws.cell(row=row, column=col)
            cell.value = f"=SUM({col_letter}{data_start}:{col_letter}{data_end})"
            cell.number_format = fmt
    style_total_row(ws, row, num_cols)
    return row + 1


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ═══════════════════════════════════════════════════════════════════
# SHEET 1: TECHNICAL BUDGET
# ═══════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Technical Budget"
ws1.sheet_properties.tabColor = "1A1A2E"
set_col_widths(ws1, [42, 16, 16, 58])

ws1.merge_cells("A1:D1")
ws1.cell(row=1, column=1).value = "AVRAI — Technical Budget"
ws1.cell(row=1, column=1).font = TITLE_FONT

ws1.merge_cells("A2:D2")
ws1.cell(row=2, column=1).value = "February 2026  •  Solo Founder  •  Pre-Pre-Seed  •  iOS + Android  •  avrai.org"
ws1.cell(row=2, column=1).font = SUBTITLE_FONT

row = 4

# ── Fixed: Subscriptions ─────────────────────────────────────────
style_section_row(ws1, row, 4, "FIXED COSTS — Subscriptions")
row += 1
row = write_headers(ws1, row, ["Item", "Monthly", "Annual", "Notes"])
data_start = row

row = write_row(ws1, row, ["Supabase Pro", 25, 300, "8GB database, 100GB storage, 250GB bandwidth, 2M edge function invocations/mo. 15+ edge functions deployed. Overage: $2/1M additional invocations."], money_cols={2, 3})
row = write_row(ws1, row, ["Cursor Ultra", 200, 2400, "AI-assisted development. 20x usage vs Pro. Unlimited model access."], money_cols={2, 3})
row = write_row(ws1, row, ["Cursor Extra Usage Credits", 3500, 42000, "Additional API usage beyond Ultra plan. Enables unrestricted model access (Opus, GPT-4, etc.)."], money_cols={2, 3})
row = write_row(ws1, row, ["Cursor Bugbot (GitHub code checking)", 40, 480, "Automated PR/code checking integration through Cursor on GitHub."], money_cols={2, 3})
row = write_row(ws1, row, ["GitHub Pro", 4, 48, "Private repos + GitHub Actions CI/CD"], money_cols={2, 3})
row = write_row(ws1, row, ["Apple Developer Account", 8.25, 99, "iOS distribution + TestFlight. $99/yr billed annually."], money_cols={2, 3}, fmt=CURRENCY_FMT_2)
data_end = row - 1
row = write_total_row(ws1, row, "TOTAL", 4, sum_cols={2, 3}, data_start=data_start, data_end=data_end, fmt=CURRENCY_FMT_2)

row += 1

# ── Fixed: One-Time Fees ─────────────────────────────────────────
style_section_row(ws1, row, 4, "FIXED COSTS — One-Time Fees")
row += 1
row = write_headers(ws1, row, ["Item", "Cost", "", "Notes"])
row = write_row(ws1, row, ["Google Play Developer", 25, None, "One-time. Required for Android distribution."], money_cols={2})

row += 1

# ── Fixed: Hardware ──────────────────────────────────────────────
style_section_row(ws1, row, 4, "FIXED COSTS — Hardware (One-Time, Buy as Needed)")
row += 1
row = write_headers(ws1, row, ["Item", "Low", "High", "Notes"])
data_start = row

row = write_row(ws1, row, ["MacBook Pro (M3/M4)", 1999, 3499, "Required for iOS builds (Xcode/macOS only). Skip if already owned."], money_cols={2, 3})
row = write_row(ws1, row, ["Mac Mini (CI server)", 599, 1299, "Automated builds/tests. Skip if using GitHub Actions."], money_cols={2, 3})
row = write_row(ws1, row, ["External Monitor (27\" 4K)", 299, 799, "Skip if already owned."], money_cols={2, 3})
row = write_row(ws1, row, ["External SSD (1-4TB)", 79, 299, "Backups, ML model files, Xcode archives"], money_cols={2, 3})
row = write_row(ws1, row, ["USB-C Hub/Dock + Cables", 99, 350, ""], money_cols={2, 3})
row = write_row(ws1, row, ["iPhone 15 Pro", 999, 1199, "Primary iOS test device. Tier 3 (full world model + SLM capable)."], money_cols={2, 3})
row = write_row(ws1, row, ["iPhone 14 (or older)", 500, 799, "Backward compatibility / Tier 2 testing. Refurbished OK."], money_cols={2, 3})
row = write_row(ws1, row, ["Google Pixel 8", 549, 699, "Primary Android test device. Stock Android baseline. Tier 2-3."], money_cols={2, 3})
row = write_row(ws1, row, ["Samsung Galaxy S24", 699, 899, "Samsung One UI behaves differently for BLE, notifications, permissions. Refurb OK."], money_cols={2, 3})
row = write_row(ws1, row, ["Budget Android (~$150 range)", 149, 299, "Low-end / Tier 0-1 performance testing. Real users have cheap phones."], money_cols={2, 3})
data_end = row - 1
row = write_total_row(ws1, row, "TOTAL", 4, sum_cols={2, 3}, data_start=data_start, data_end=data_end)

row += 1

# ── Variable: APIs & Cloud ───────────────────────────────────────
style_section_row(ws1, row, 4, "VARIABLE COSTS — APIs & Cloud Services")
row += 1
row = write_headers(ws1, row, ["Item", "Current Cost", "What It Does", "Notes"])

row = write_row(ws1, row, ["Google Gemini API", 0, "Powers AI chat via llm-chat-stream edge function", "Free tier: 1,000 req/day, 15 RPM. Paid: ~$1.25/1M input tokens. Gemini Flash cuts 75%."], money_cols={2})
row = write_row(ws1, row, ["Google Places API (New)", 0, "Text search, nearby search, place details for spot discovery", "$200/mo free credit. $17/1K text searches. Credit covers ~11K searches/mo."], money_cols={2})
row = write_row(ws1, row, ["Google Maps SDK", 0, "Renders native map view", "$200/mo free credit shared with Places. $7/1K map loads."], money_cols={2})
row = write_row(ws1, row, ["Firebase (Blaze plan)", 0, "Auth, Firestore, Storage (tax docs), Analytics (free), FCM push (free)", "Free tier: 50K reads/day, 20K writes/day, 5GB storage."], money_cols={2})
row = write_row(ws1, row, ["OpenWeatherMap", 0, "Weather context for recommendations (indoor spots on rainy days)", "Free tier: 1,000 calls/day. API key not yet configured."], money_cols={2})
row = write_row(ws1, row, ["Stripe", 0, "Payments, refunds, identity verification", "No monthly fee. 2.9% + $0.30/txn. $1.50/identity verification. Costs only when processing real money."], money_cols={2})
row = write_row(ws1, row, ["Google Cloud Storage", 0, "File/media storage beyond Supabase", "5GB free. $0.020/GB/mo after."], money_cols={2})
row = write_row(ws1, row, ["Supabase Storage (model distribution)", 0, "On-device ONNX models (~1MB) + optional SLM (700MB-2GB) downloaded by users", "Included in Pro plan. Bandwidth = $0.09/GB egress. Migrate to Cloudflare R2 (free egress) before costs grow."], money_cols={2})
row = write_row(ws1, row, ["Cloudflare R2 (future CDN)", 0, "Model delivery with free egress", "Not yet set up. $0.015/GB/mo storage. Use when model downloads create real bandwidth cost."], money_cols={2})
row = write_row(ws1, row, ["ML Training Compute (Cloud GPU)", 0, "Pre-training energy function, transition predictor, batch retraining", "Google Colab free tier for now. RunPod/Lambda $10-$50/run when needed."], money_cols={2})

row += 1

# ── Planned Resilience/Governance Reserve ────────────────────────
style_section_row(ws1, row, 4, "PLANNED RESILIENCE/GOVERNANCE RESERVE (Master Plan 10.9.x + ML Governance)")
row += 1
row = write_headers(ws1, row, ["Item", "Monthly (Low)", "Monthly (High)", "Why It Exists"])
data_start = row
row = write_row(ws1, row, ["Observability + reliability tooling reserve", 50, 300, "Dashboards/alerts and operational visibility for break-to-learning metrics (TTD, TTH, recurrence, impact radius)."], money_cols={2, 3}, fmt=CURRENCY_FMT_2)
row = write_row(ws1, row, ["GitHub CI/check overage reserve", 20, 150, "Additional required checks/workflow volume from execution + traceability + architecture + ML governance guards."], money_cols={2, 3}, fmt=CURRENCY_FMT_2)
row = write_row(ws1, row, ["ML retraining/simulation compute reserve", 100, 600, "Non-free burst training/simulation runs and replay cycles as model scope expands."], money_cols={2, 3}, fmt=CURRENCY_FMT_2)
row = write_row(ws1, row, ["Data pipeline/storage/egress reserve", 25, 200, "Training artifacts, experiment logs, recovery queue growth, and transfer overhead."], money_cols={2, 3}, fmt=CURRENCY_FMT_2)
row = write_row(ws1, row, ["Contingency on resilience reserve (10-15%)", 19.5, 187.5, "Buffer for reopen events and unexpected remediation loops."], money_cols={2, 3}, fmt=CURRENCY_FMT_2)
data_end = row - 1
row = write_total_row(ws1, row, "TOTAL RESILIENCE RESERVE ENVELOPE", 4, sum_cols={2, 3}, data_start=data_start, data_end=data_end, fmt=CURRENCY_FMT_2)

row += 1

# ── Free Services ────────────────────────────────────────────────
style_section_row(ws1, row, 4, "FREE SERVICES — No Budget Impact")
row += 1
row = write_headers(ws1, row, ["Item", "What It Does", "", "Why Free"])

row = write_row(ws1, row, ["OpenStreetMap (Nominatim + Overpass)", "Place search fallback, nearby amenity queries", "", "Community API, free with rate limiting"])
row = write_row(ws1, row, ["Apple MapKit (MKLocalSearch)", "Native iOS place search with cache", "", "Native SDK, no per-call cost"])
row = write_row(ws1, row, ["Firebase Analytics", "Event tracking", "", "Always free, unlimited"])
row = write_row(ws1, row, ["Firebase Cloud Messaging", "Push notifications", "", "Always free, unlimited"])
row = write_row(ws1, row, ["Google OAuth", "Google Sign-In", "", "Free"])
row = write_row(ws1, row, ["Embeddings edge function", "Text embeddings", "", "Uses local hash, no external API"])
row = write_row(ws1, row, ["Social media OAuth (10 platforms)", "Instagram, Facebook, X, Reddit, TikTok, Pinterest, LinkedIn, YouTube, Tumblr, Are.na", "", "OAuth endpoints free. Behind feature flag."])
row = write_row(ws1, row, ["HuggingFace", "Token configured, no active calls", "", "Free tier"])
row = write_row(ws1, row, ["TestFlight", "iOS beta distribution", "", "Included with Apple Developer"])
row = write_row(ws1, row, ["Signal Protocol", "End-to-end encryption for AI2AI", "", "Runs entirely on-device (Rust FFI)"])
row = write_row(ws1, row, ["ONNX Runtime", "On-device world model inference (energy function, state encoder, transition predictor, action encoder)", "", "Open-source, bundled in app binary"])
row = write_row(ws1, row, ["Drift / SQLite", "On-device episodic memory, semantic memory storage", "", "Open-source, on-device"])
row = write_row(ws1, row, ["PyTorch + ONNX tools", "Training pipeline for world model components", "", "Open-source Python libraries"])


# ═══════════════════════════════════════════════════════════════════
# SHEET 2: OPERATIONAL BUDGET
# ═══════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Operational Budget")
ws2.sheet_properties.tabColor = "283593"
set_col_widths(ws2, [44, 16, 16, 58])

ws2.merge_cells("A1:D1")
ws2.cell(row=1, column=1).value = "AVRAI — Operational Budget"
ws2.cell(row=1, column=1).font = TITLE_FONT

ws2.merge_cells("A2:D2")
ws2.cell(row=2, column=1).value = "February 2026  •  RG Enterprises LLC (Alabama) + AVRAI LLC (Delaware)"
ws2.cell(row=2, column=1).font = SUBTITLE_FONT

row = 4

# ── Business Formation ───────────────────────────────────────────
style_section_row(ws2, row, 4, "FIXED COSTS — Business Formation (One-Time)")
row += 1
row = write_headers(ws2, row, ["Item", "Low", "High", "Notes"])
data_start = row

row = write_row(ws2, row, ["Alabama LLC Filing (RG Enterprises)", 200, 200, "AL Secretary of State filing fee"], money_cols={2, 3})
row = write_row(ws2, row, ["Delaware LLC Filing (AVRAI)", 90, 90, "DE Division of Corporations filing fee"], money_cols={2, 3})
row = write_row(ws2, row, ["Delaware Registered Agent (AVRAI) — first year", 50, 300, "Required for DE LLC. Annual recurring cost (see Recurring section)."], money_cols={2, 3})
row = write_row(ws2, row, ["Alabama Registered Agent (RG Enterprises) — first year", 0, 100, "Can serve as your own if you have an AL address. Otherwise $50-$100/yr."], money_cols={2, 3})
row = write_row(ws2, row, ["Operating Agreement — RG Enterprises", 500, 1500, "Attorney-drafted. Ownership, equity, voting."], money_cols={2, 3})
row = write_row(ws2, row, ["Operating Agreement — AVRAI", 500, 1500, "Attorney-drafted. Defines relationship to RG Enterprises if parent/holding entity."], money_cols={2, 3})
row = write_row(ws2, row, ["EIN Registration — RG Enterprises", 0, 0, "Free from IRS"], money_cols={2, 3})
row = write_row(ws2, row, ["EIN Registration — AVRAI", 0, 0, "Free from IRS"], money_cols={2, 3})
row = write_row(ws2, row, ["Business License (Alabama)", 50, 300, "Varies by city/county. Some jurisdictions don't require for software."], money_cols={2, 3})
row = write_row(ws2, row, ["Privacy Policy (attorney-drafted)", 500, 2000, "GDPR + CCPA compliance already coded. Legal doc needs to match implementation."], money_cols={2, 3})
row = write_row(ws2, row, ["Terms of Service (attorney-drafted)", 500, 2000, "Already linked from onboarding. Needs real legal language."], money_cols={2, 3})
data_end = row - 1
row = write_total_row(ws2, row, "TOTAL", 4, sum_cols={2, 3}, data_start=data_start, data_end=data_end)

row += 1

# ── Recurring Operational ────────────────────────────────────────
style_section_row(ws2, row, 4, "FIXED COSTS — Recurring Operational")
row += 1
row = write_headers(ws2, row, ["Item", "Monthly", "Annual", "Notes"])
data_start = row

row = write_row(ws2, row, ["Google Workspace (Business Starter)", 7, 84, "Custom email @avrai.org (support@, privacy@, feedback@, bugs@, business-support@, demo@). Promo: $3.50/mo through May 2026."], money_cols={2, 3})
row = write_row(ws2, row, ["Claude Pro (or equivalent AI)", 20, 240, "Admin, legal research, operational tasks, document drafting"], money_cols={2, 3})
row = write_row(ws2, row, ["avrai.org Domain Renewal", None, 12, "Already owned"], money_cols={3})
row = write_row(ws2, row, ["Delaware Registered Agent (AVRAI)", None, 150, "Required annually for DE LLC. Range: $50-$300/yr. Midpoint used."], money_cols={3})
row = write_row(ws2, row, ["Alabama Registered Agent (RG Enterprises)", None, 0, "Free if you serve as your own agent with an AL address. Otherwise $50-$100/yr."], money_cols={3})
row = write_row(ws2, row, ["Accounting Software", 0, 0, "Wave (free) or QuickBooks ($30/mo). Only needed once real transactions exist. Using $0 until then."], money_cols={2, 3})
data_end = row - 1
row = write_total_row(ws2, row, "TOTAL", 4, sum_cols={2, 3}, data_start=data_start, data_end=data_end)

row += 1

# ── Design ───────────────────────────────────────────────────────
style_section_row(ws2, row, 4, "VARIABLE COSTS — Design (One-Time)")
row += 1
row = write_headers(ws2, row, ["Item", "Low", "High", "Notes"])
data_start = row

row = write_row(ws2, row, ["App Designer (UI/UX contract)", 3000, 10000, "Full design system: all core screens, component library, interaction patterns, accessibility. Low = experienced freelancer. High = boutique agency."], money_cols={2, 3})
row = write_row(ws2, row, ["Logo + App Icon", 200, 2000, "Fiverr ($200-500) to professional brand designer ($1,500+). Must work at all sizes (1024x1024 down to 29x29)."], money_cols={2, 3})
row = write_row(ws2, row, ["Design Tools (Figma)", 0, 0, "Free for individual use. Designer uses their own account."], money_cols={2, 3})
data_end = row - 1
row = write_total_row(ws2, row, "TOTAL", 4, sum_cols={2, 3}, data_start=data_start, data_end=data_end)

row += 1

# ── IP ───────────────────────────────────────────────────────────
style_section_row(ws2, row, 4, "VARIABLE COSTS — Intellectual Property (One-Time)")
row += 1
row = write_headers(ws2, row, ["Item", "Low", "High", "Notes"])
data_start = row

row = write_row(ws2, row, ["Patent Attorney Consultation", 2000, 5000, "Strategy session + prior art search across all patent areas"], money_cols={2, 3})
row = write_row(ws2, row, ["Provisional #1: AI2AI Personality Learning", 2000, 5000, "Core tech. AI agents learn/evolve personality via device-to-device communication. Most defensible IP."], money_cols={2, 3})
row = write_row(ws2, row, ["Provisional #2: Quantum-Inspired Matching", 2000, 5000, "Quantum state personality representation + compatibility via inner products"], money_cols={2, 3})
row = write_row(ws2, row, ["Provisional #3: Offline-First Mesh Network", 2000, 5000, "BLE device discovery and mesh protocol for AI2AI without internet"], money_cols={2, 3})
row = write_row(ws2, row, ["Provisional #4: Knot Theory Group Dynamics", 2000, 5000, "Group formation using topological knot invariants"], money_cols={2, 3})
row = write_row(ws2, row, ["Provisional #5: Privacy-Preserving Architecture", 2000, 5000, "Differential privacy, k-anonymity, internal/external boundary"], money_cols={2, 3})
data_end = row - 1
row = write_total_row(ws2, row, "TOTAL (Provisionals)", 4, sum_cols={2, 3}, data_start=data_start, data_end=data_end)

row = write_row(ws2, row, ["Non-Provisional Conversion (per patent)", 7040, 23080, "12-month window from provisional filing date. All 5 = $35,200-$115,400. Defer to post-seed."], money_cols={2, 3})

row += 1

# ── Legal/Compliance ─────────────────────────────────────────────
style_section_row(ws2, row, 4, "VARIABLE COSTS — Legal & Compliance (One-Time)")
row += 1
row = write_headers(ws2, row, ["Item", "Low", "High", "Notes"])
data_start = row

row = write_row(ws2, row, ["Cryptographic Security Audit", 15000, 30000, "Signal Protocol, AI2AI encryption, BLE security, penetration testing. Required before handling real user data."], money_cols={2, 3})
row = write_row(ws2, row, ["GDPR Compliance Review", 0, 3000, "Only if targeting EU at launch. Code-level compliance already built."], money_cols={2, 3})
row = write_row(ws2, row, ["Ongoing Legal Consultation", 0, 2000, "Contract review, NDAs. As needed. Per quarter."], money_cols={2, 3})
data_end = row - 1
row = write_total_row(ws2, row, "TOTAL", 4, sum_cols={2, 3}, data_start=data_start, data_end=data_end)


# ═══════════════════════════════════════════════════════════════════
# SHEET 3: SUMMARY
# ═══════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Summary")
ws3.sheet_properties.tabColor = "4CAF50"
set_col_widths(ws3, [46, 18, 18, 48])

ws3.merge_cells("A1:D1")
ws3.cell(row=1, column=1).value = "AVRAI — Budget Summary"
ws3.cell(row=1, column=1).font = TITLE_FONT

ws3.merge_cells("A2:D2")
ws3.cell(row=2, column=1).value = "February 2026  •  Solo Founder  •  Pre-Pre-Seed"
ws3.cell(row=2, column=1).font = SUBTITLE_FONT

row = 4

# ── Monthly Burn ─────────────────────────────────────────────────
style_section_row(ws3, row, 4, "MONTHLY BURN — Development Phase")
row += 1
row = write_headers(ws3, row, ["Item", "Monthly", "", "Notes"])
data_start = row

row = write_row(ws3, row, ["Supabase Pro", 25, None, "Database + storage + edge functions"], money_cols={2})
row = write_row(ws3, row, ["Cursor Ultra", 200, None, "AI-assisted development. 20x usage vs Pro."], money_cols={2})
row = write_row(ws3, row, ["Cursor Extra Usage Credits", 3500, None, "Unrestricted model access (Opus, GPT-4, etc.)"], money_cols={2})
row = write_row(ws3, row, ["Cursor Bugbot (GitHub code checking)", 40, None, "Automated PR/code checking integration through Cursor on GitHub"], money_cols={2})
row = write_row(ws3, row, ["Claude Pro", 20, None, "Admin, research, drafting"], money_cols={2})
row = write_row(ws3, row, ["GitHub Pro", 4, None, "Private repos + CI/CD"], money_cols={2})
row = write_row(ws3, row, ["Google Workspace", 7, None, "@avrai.org email ($3.50/mo promo through May '26)"], money_cols={2})
row = write_row(ws3, row, ["Apple Developer (amortized)", 8.25, None, "$99/yr billed annually"], money_cols={2}, fmt=CURRENCY_FMT_2)
row = write_row(ws3, row, ["All APIs & Cloud Services", 0, None, "All free tiers during solo development"], money_cols={2})
data_end = row - 1
# Total row
total_row = row
ws3.cell(row=row, column=1).value = "TOTAL MONTHLY BURN"
col_letter = get_column_letter(2)
ws3.cell(row=row, column=2).value = f"=SUM({col_letter}{data_start}:{col_letter}{data_end})"
ws3.cell(row=row, column=2).number_format = CURRENCY_FMT_2
# Annualized
ws3.cell(row=row, column=3).value = f"=B{row}*12"
ws3.cell(row=row, column=3).number_format = CURRENCY_FMT
ws3.cell(row=row, column=4).value = "Per year"
style_total_row(ws3, row, 4)
row += 1

# Planned reserve envelope (separate from committed burn)
row = write_row(ws3, row, ["Resilience reserve envelope (low-high)", 214.5, 1437.5, "Master-plan 10.9.x + ML governance reserve"], money_cols={2, 3}, fmt=CURRENCY_FMT_2)
row = write_row(ws3, row, ["Total monthly burn with reserve", 4018.75, 5241.75, "Committed burn + resilience reserve envelope"], money_cols={2, 3}, fmt=CURRENCY_FMT_2)
row = write_row(ws3, row, ["Total annual burn with reserve", 48225, 62901, "Annualized from monthly reserve-inclusive range"], money_cols={2, 3})

row += 1

# ── One-time by priority ────────────────────────────────────────
style_section_row(ws3, row, 4, "ONE-TIME COSTS — By Priority")
row += 1
row = write_headers(ws3, row, ["Item", "Low", "High", "When"])

row = write_row(ws3, row, ["Legal formation (2 LLCs + legal docs)", 2390, 7990, "Now"], money_cols={2, 3})
row = write_row(ws3, row, ["Google Play Developer account", 25, 25, "Now"], money_cols={2, 3})
row = write_row(ws3, row, ["Test devices (1-5 phones)", 149, 3895, "Before beta"], money_cols={2, 3})
row = write_row(ws3, row, ["App designer (UI/UX + logo)", 3200, 12000, "Before beta"], money_cols={2, 3})
row = write_row(ws3, row, ["Development hardware (if needed)", 3075, 6246, "As needed — skip if already owned"], money_cols={2, 3})
row = write_row(ws3, row, ["Patent attorney + 5 provisionals", 12000, 30000, "When fundable"], money_cols={2, 3})
row = write_row(ws3, row, ["Security audit + legal compliance", 15000, 35000, "Post-raise"], money_cols={2, 3})

row += 1

# ── Total Capital Needed ─────────────────────────────────────────
style_section_row(ws3, row, 4, "TOTAL CAPITAL NEEDED — Scenarios (Year 1)")
row += 1
row = write_headers(ws3, row, ["Scenario", "One-Time", "Year 1 Burn", "Year 1 Total"])

    # Minimum: burn + 2 LLC filings + EINs + 1 cheap phone
row = write_row(ws3, row, ["Minimum viable (burn + LLCs + 1 phone)", 2564, 45651, 48215], money_cols={2, 3, 4})
# Comfortable: above + designer low + 3 phones + legal docs
row = write_row(ws3, row, ["Comfortable to beta (+ designer + phones + legal docs)", 10860, 45651, 56511], money_cols={2, 3, 4})
# Full pre-seed: above + patents
row = write_row(ws3, row, ["Full pre-seed (+ patents)", 22860, 45651, 68511], money_cols={2, 3, 4})
# Everything: above + security audit + all hardware
row = write_row(ws3, row, ["Everything (+ security audit + all hardware)", 69306, 45651, 114957], money_cols={2, 3, 4})


# ── Save ─────────────────────────────────────────────────────────
output_path = "/Users/reisgordon/AVRAI/docs/business/AVRAI_BUDGET.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
